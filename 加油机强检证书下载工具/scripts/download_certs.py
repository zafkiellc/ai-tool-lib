#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
加油站强制检定证书批量下载主脚本
=================================
流程（已实测验证）：
1. 打开 http://scjg.hubei.gov.cn/hbjl/a/login
2. 点击顶部"登录" -> 点击"统一身份认证登录" -> 法人登录
3. 填统一社会信用代码 + 密码（无需验证码），点击登录
4. 直达列表 URL: http://scjg.hubei.gov.cn/hbjl/a#/asset/strongcheck/queryForce
5. 遍历每行 -> 点"查看" -> 详情页提取枪号+证书编号 -> 点可见的证书预览图标下载PDF
6. 保存为 D:/加油机强检证书/<站点名>/枪<枪号>_<证书编号>.pdf

运行：
  C:/Users/zafki/.workbuddy/binaries/python/versions/3.13.12/python.exe download_certs.py
"""
import json
import re
import sys
import time
from pathlib import Path
from datetime import datetime

import openpyxl
from playwright.sync_api import sync_playwright

# ================= 配置 =================
EXCEL_PATH = Path("C:/Users/zafki/OneDrive/工作/安数/账号/加油站强制检定申报信息统计表.xlsx")
OUTPUT_DIR = Path("D:/加油机强检证书")
LOG_FILE = OUTPUT_DIR / "download_log.json"
LIST_URL = "http://scjg.hubei.gov.cn/hbjl/a#/asset/strongcheck/queryForce"
LOGIN_URL = "http://scjg.hubei.gov.cn/hbjl/a/login"

# 浏览器路径：优先便携版（PLAYWRIGHT_BROWSERS_PATH），否则用本机开发环境
import os as _os
_browsers_root = _os.environ.get("PLAYWRIGHT_BROWSERS_PATH") or str(Path.home() / "AppData/Local/ms-playwright")
CHROME_EXE = str(Path(_browsers_root) / "chromium-1234" / "chrome-win64" / "chrome.exe")

# 只下载指定站点（None = 全部）。例：["捷达", "十升"]
ONLY_STATIONS = None
# 跳过已成功的站点（断点续传）
SKIP_DONE = True

# ================= 工具函数 =================

def clean_filename(text):
    """清理文件名非法字符。"""
    if text is None:
        return "unknown"
    return re.sub(r'[\\/:*?"<>|\s]+', "_", str(text).strip()) or "unknown"


def read_stations(excel_path):
    """读取加油站清单。"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    stations = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        station = row[1]
        username = row[2]
        password = row[3]
        guns = row[7]
        if not station or not username:
            continue
        stations.append({
            "name": str(station).strip(),
            "username": str(username).strip(),
            "password": str(password).strip() if password else "",
            "guns": guns,
        })
    wb.close()
    return stations


def wait_url_contains(page, substr, timeout=60):
    """轮询等待 URL 包含指定子串（避免 networkidle 因 WebSocket 卡死）。"""
    deadline = time.time() + timeout
    while time.time() < deadline:
        if substr in page.url:
            return True
        page.wait_for_timeout(1000)
    raise Exception(f"等待URL包含 '{substr}' 超时, 当前URL: {page.url}")


def open_login_page(page, max_wait=600):
    """
    打开登录页并等待可交互。登录页可能被 WAF 间歇性限流（412/空白），
    轮询刷新直到出现"登录"按钮，最多等 max_wait 秒。
    """
    deadline = time.time() + max_wait
    while time.time() < deadline:
        try:
            page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=45000)
        except Exception:
            pass
        page.wait_for_timeout(3000)
        if page.get_by_text("登录", exact=True).count() > 0:
            page.wait_for_timeout(1500)
            return True
        print(f"  登录页暂不可用，15 秒后重试... ({int(deadline - time.time())}s 剩余)")
        page.wait_for_timeout(15000)
    return False


def login(page, username, password, max_retries=3):
    """统一身份认证 - 法人登录（已实测）。带等待与重试。"""
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            if attempt > 1:
                print(f"  登录重试 {attempt}/{max_retries}，冷却 300 秒...")
                time.sleep(300)  # 风控冷却
            # 打开登录页（轮询直到可交互）
            if not open_login_page(page):
                raise Exception("登录页长时间不可访问（可能被限流）")
            # 等"登录"链接可见
            login_link = page.get_by_text("登录", exact=True).first
            login_link.wait_for(state="visible", timeout=30000)
            login_link.click(no_wait_after=True)
            page.wait_for_timeout(1500)
            # 统一身份认证登录
            auth_link = page.get_by_text("统一身份认证登录", exact=False).first
            auth_link.wait_for(state="visible", timeout=30000)
            auth_link.click(no_wait_after=True)
            # 等待跳到统一身份认证页（oauth.hubei.gov.cn）
            wait_url_contains(page, "oauth.hubei.gov.cn", timeout=60)
            page.wait_for_timeout(2500)
            # 法人登录 tab
            fr_tab = page.locator("#tab_fr")
            fr_tab.wait_for(state="visible", timeout=30000)
            fr_tab.click(no_wait_after=True)
            page.wait_for_timeout(1200)
            # 填账号密码
            page.locator("#loginName_ent").fill(username)
            page.locator("#password_ent").fill(password)
            page.locator("#loginform .btns[onclick*='uiasEntLogin']").click(no_wait_after=True)
            # 等待跳回原系统（hbjl/a#/home 或含 hbjl 的页面）
            wait_url_contains(page, "scjg.hubei.gov.cn/hbjl", timeout=90)
            page.wait_for_timeout(4000)
            # 校验登录成功（出现个人中心/安全退出）
            try:
                page.get_by_text("安全退出", exact=False).first.wait_for(state="visible", timeout=15000)
            except Exception:
                pass
            if "hbjl/a#/home" in page.url or "hbjl" in page.url:
                return True
            else:
                raise Exception(f"登录后URL异常: {page.url}")
        except Exception as e:
            last_err = e
            print(f"  登录尝试 {attempt} 失败: {e}")
            page.wait_for_timeout(3000)
    raise Exception(f"登录失败（重试{max_retries}次）: {last_err}")


def extract_cert_info(page):
    """
    从详情页提取枪号、证书编号。
    返回 (gun_no, cert_no, has_icon)。
    """
    result = page.evaluate('''() => {
        // 1. 找可见的证书预览图标
        const icons = Array.from(document.querySelectorAll('i[title="证书预览"]'));
        let visibleIcon = null;
        for (const i of icons) {
            const r = i.getBoundingClientRect();
            const s = window.getComputedStyle(i);
            if (r.width > 0 && r.height > 0 && s.display !== 'none') {
                visibleIcon = i;
                break;
            }
        }
        if (!visibleIcon) return {gunNo: '', certNo: '', hasIcon: false};

        // 2. 向上找容器文本（含证书编号）
        let container = visibleIcon.closest('.row') || visibleIcon.parentElement;
        let containerText = '';
        for (let j=0; j<4 && container; j++) {
            if (container.innerText && container.innerText.trim()) {
                containerText = container.innerText.trim();
                break;
            }
            container = container.parentElement;
        }

        // 3. 提取证书编号
        let certNo = '';
        const m = containerText.match(/上次检定证书编号[\\s\\n]*([^\\n]+)/);
        if (m) certNo = m[1].trim();

        // 4. 找枪号：匹配 "枪号" 或 "*枪号" 的label（星号可能在子元素里）
        let gunNo = '';
        const labels = Array.from(document.querySelectorAll('.control-label, label, th, span, div'));
        const gunLabel = labels.find(e => {
            const t = (e.textContent || '').trim();
            const plain = t.replace(/[\\*\\s]/g, '');
            return plain === '枪号' || t === '*枪号';
        });
        if (gunLabel) {
            // 直接取gunLabel所在行的值：找同一个 .row 下的 value div
            let row = gunLabel.parentElement;
            for (let i=0; i<5 && row; i++) {
                const t = row.innerText || '';
                const gm = t.match(/[\\*]?枪号[\\s\\n]*([^\\n]+)/);
                if (gm && gm[1].trim() && !gm[1].trim().includes('如果器具')) {
                    gunNo = gm[1].trim();
                    break;
                }
                row = row.parentElement;
            }
        }
        return {gunNo, certNo, hasIcon: true};
    }''')
    return result.get("gunNo", ""), result.get("certNo", ""), result.get("hasIcon", False)


def download_cert(page, context, save_path, timeout_ms=20000):
    """点击可见的证书预览图标并保存下载。"""
    icon = page.locator("i[title='证书预览']").locator("visible=true").first
    # 等图标可见
    icon.wait_for(state="visible", timeout=10000)
    with page.expect_download(timeout=timeout_ms) as dl_info:
        icon.click(no_wait_after=True)
    download = dl_info.value
    download.save_as(str(save_path))
    return download.suggested_filename


# 停止标志：web_launcher 可设置，process_station 每把枪前检查
STOP_REQUESTED = False


def stop_check():
    """是否收到停止请求。"""
    global STOP_REQUESTED
    return STOP_REQUESTED


def launch_browser(p):
    """
    启动浏览器：直接调用系统 Edge/Chrome（WAF 指纹检测可通过），
    移除 --enable-automation 标志（否则湖北省强检系统返回 412）。
    不再依赖 playwright 自带 chromium（可清理 browsers/ 目录省空间）。
    """
    common = dict(
        headless=True,
        ignore_default_args=["--enable-automation"],
        args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
    )
    last_err = None
    for channel in ("msedge", "chrome"):
        try:
            print(f"  尝试系统浏览器: {channel}")
            return p.chromium.launch(channel=channel, **common)
        except Exception as e:
            last_err = e
            print(f"  {channel} 不可用: {e}")
    raise RuntimeError(
        f"未找到可用的系统浏览器（Edge/Chrome）: {last_err}\n"
        "请安装 Microsoft Edge 或 Google Chrome 后重试。"
    )


def process_station(page, context, station, output_dir, log):
    """处理单个站点全部证书下载。"""
    name = station["name"]
    station_dir = output_dir / clean_filename(name)
    station_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n{'='*70}")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 站点: {name} (账号: {station['username']})")

    # 登录（每个站点是全新浏览器实例；登录前随机等待 15-30 秒）
    import random as _random
    wait_sec = _random.randint(15, 30)
    # 等待期间可停止：分片 sleep 并检查停止标志
    print(f"  等待 {wait_sec} 秒后开始登录（防风控）...")
    for _i in range(wait_sec):
        if stop_check():
            print(f"  ! 收到停止请求，取消登录等待")
            log[name] = {"status": "stopped", "downloaded": 0, "total": 0, "failed": [], "dir": str(station_dir),
                         "time": datetime.now().isoformat()}
            return
        time.sleep(1)
    try:
        login(page, station["username"], station["password"])
        print(f"  登录成功: {page.url}")
    except Exception as e:
        print(f"  登录失败: {e}")
        log[name] = {"status": "login_failed", "error": str(e), "time": datetime.now().isoformat()}
        return

    # 进入列表
    page.goto(LIST_URL, wait_until="domcontentloaded")
    # 等待表格出现
    try:
        page.locator("table tbody tr").first.wait_for(state="visible", timeout=30000)
    except Exception:
        print("  ! 列表表格未出现，等待后重试...")
        page.wait_for_timeout(8000)
    page.wait_for_timeout(1500)

    downloaded = 0
    failed_rows = []
    processed_codes = set()   # 已处理的赋码（防重复）
    row_index = 0

    while True:
        rows = page.locator("table tbody tr").all()
        # 只保留含"查看"的行
        rows = [r for r in rows if r.locator("a:has-text('查看')").count() > 0]
        if not rows:
            break

        for row in rows:
            # 停止检查：每把枪前
            if stop_check():
                print(f"  ! 收到停止请求，提前结束本站点（已下载 {downloaded} 份）")
                break

            # 取赋码（td[1]）作为唯一标识
            try:
                code_cell = row.locator("td").nth(1)
                fuma = code_cell.inner_text().strip()
            except Exception:
                fuma = ""

            if fuma and fuma in processed_codes:
                print(f"  - 跳过已处理赋码: {fuma}")
                continue

            row_index += 1
            try:
                cells = row.locator("td").all()
                row_desc = " | ".join(c.inner_text().strip()[:40] for c in cells[:6])
                print(f"  [{row_index}] {row_desc}")

                # 点击查看
                row.locator("a:has-text('查看')").first.click(no_wait_after=True)
                # 等待详情页加载（出现证书预览图标或 mainContainer）
                try:
                    page.locator("i[title='证书预览']").first.wait_for(state="attached", timeout=25000)
                except Exception:
                    page.wait_for_timeout(6000)
                page.wait_for_timeout(2000)

                # 提取枪号、证书编号
                gun_no, cert_no, has_icon = extract_cert_info(page)

                if not has_icon:
                    print(f"    - 无可见证书预览图标（可能无证书），跳过")
                    failed_rows.append({"row": row_index, "reason": "no_cert_icon"})
                    processed_codes.add(fuma)
                else:
                    # 文件名
                    gun_part = f"枪{clean_filename(gun_no)}" if gun_no else f"行{row_index}"
                    cert_part = clean_filename(cert_no) if cert_no else f"no_cert_{row_index}"
                    save_path = station_dir / f"{gun_part}_{cert_part}.pdf"

                    # 下载
                    try:
                        suggested = download_cert(page, context, save_path)
                        print(f"    ✓ 已下载: {save_path.name} (源: {suggested})")
                        downloaded += 1
                        processed_codes.add(fuma)
                    except Exception as e:
                        print(f"    ! 下载失败: {e}")
                        failed_rows.append({"row": row_index, "cert": cert_no, "error": str(e)})
                        processed_codes.add(fuma)
                        try:
                            page.screenshot(path=str(station_dir / f"error_row{row_index}.png"))
                        except Exception:
                            pass

            except Exception as e:
                print(f"    ! 处理失败: {e}")
                failed_rows.append({"row": row_index, "error": str(e)})
                if fuma:
                    processed_codes.add(fuma)

            # 返回列表：优先 go_back 保留分页状态
            try:
                page.go_back(wait_until="domcontentloaded")
                page.wait_for_timeout(2500)
            except Exception:
                try:
                    page.goto(LIST_URL, wait_until="domcontentloaded")
                    page.wait_for_timeout(2500)
                except Exception:
                    pass

        # 分页：点击"下一页"后对比首行赋码，若未变化则停止
        next_btn = page.locator("a:has-text('下一页')").first
        if next_btn.count() > 0:
            try:
                # 当前首行赋码
                cur_rows = page.locator("table tbody tr").all()
                cur_first_code = ""
                for r in cur_rows:
                    if r.locator("a:has-text('查看')").count() > 0:
                        cur_first_code = r.locator("td").nth(1).inner_text().strip()
                        break
                # 检查是否disabled
                li_cls = next_btn.locator("xpath=..").get_attribute("class") or ""
                if "disabled" in li_cls:
                    break
                next_btn.click(no_wait_after=True)
                page.wait_for_timeout(3500)
                # 点击后首行赋码
                new_rows = page.locator("table tbody tr").all()
                new_first_code = ""
                for r in new_rows:
                    if r.locator("a:has-text('查看')").count() > 0:
                        new_first_code = r.locator("td").nth(1).inner_text().strip()
                        break
                if new_first_code and new_first_code == cur_first_code:
                    print("  分页无变化，停止")
                    break
                continue
            except Exception:
                break
        else:
            break

    print(f"  站点 {name} 完成: 成功 {downloaded} 份, 失败 {len(failed_rows)} 条")
    log[name] = {
        "status": "done",
        "downloaded": downloaded,
        "total": row_index,
        "failed": failed_rows,
        "dir": str(station_dir),
        "time": datetime.now().isoformat(),
    }


def main():
    if not EXCEL_PATH.exists():
        print(f"错误: Excel 不存在: {EXCEL_PATH}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stations = read_stations(EXCEL_PATH)
    print(f"读取到 {len(stations)} 个加油站")

    if ONLY_STATIONS:
        stations = [s for s in stations if s["name"] in ONLY_STATIONS]
        print(f"过滤后 {len(stations)} 个站点")

    # 日志
    log = {}
    if LOG_FILE.exists():
        try:
            log = json.loads(LOG_FILE.read_text(encoding="utf-8"))
        except Exception:
            log = {}

    # 多轮重试：每轮处理所有未完成站点；每个站点用全新浏览器实例（绕 WAF 实例级限流）
    MAX_ROUNDS = 5
    for round_no in range(1, MAX_ROUNDS + 1):
        pending = [
            s for s in stations
            if not (SKIP_DONE and log.get(s["name"], {}).get("status") == "done")
        ]
        if not pending:
            print(f"\n第 {round_no} 轮：所有站点已完成")
            break
        print(f"\n{'#'*60}\n第 {round_no}/{MAX_ROUNDS} 轮，待处理 {len(pending)} 个站点\n{'#'*60}")

        for station in pending:
            if stop_check():
                print("收到停止请求，结束全部任务")
                break
            with sync_playwright() as p:
                browser = launch_browser(p)
                try:
                    context = browser.new_context(
                        viewport={"width": 1366, "height": 900},
                        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                        accept_downloads=True,
                    )
                    page = context.new_page()
                    page.set_default_timeout(30000)
                    process_station(page, context, station, OUTPUT_DIR, log)
                except Exception as e:
                    print(f"站点 {station['name']} 异常: {e}")
                    log[station["name"]] = {"status": "error", "error": str(e), "time": datetime.now().isoformat()}
                finally:
                    try:
                        browser.close()
                    except Exception:
                        pass
            # 每站保存日志
            LOG_FILE.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")

        # 检查是否还有未完成站点
        remaining = [s for s in stations if log.get(s["name"], {}).get("status") != "done"]
        if remaining and round_no < MAX_ROUNDS:
            print(f"\n还有 {len(remaining)} 个站点未完成: {[s['name'] for s in remaining]}")
            print("等待 5 分钟后开始下一轮...")
            time.sleep(300)

    print("\n" + "="*70)
    print("全部站点处理完毕!")
    print(f"日志: {LOG_FILE}")


if __name__ == "__main__":
    main()
