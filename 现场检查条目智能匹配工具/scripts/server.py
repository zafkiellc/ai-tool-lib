#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
督导系统自动填表助手 - 本地服务

启动后：
  - inspector_matcher.html (工具) → POST /api/push 推条目；标准相关接口见下
  - 360极速X 上的油猴脚本       → GET  /api/pending 拉条目
                              ←  POST /api/remove 填完移除

待录入队列数据存 pending.json（同目录）。
上传的检查标准存 standards/ 目录（每个标准一个 json + standards_index.json）。
CORS 全开放。

运行形态：
  - python server.py              → 控制台运行（开发用）
  - server.exe (--noconsole)      → 后台运行，弹出系统托盘图标，右键菜单「打开网页/退出」
"""
import http.server
import json
import os
import sys
import time
import uuid
import threading
import base64
import io
import webbrowser
from urllib.parse import urlparse

# 单文件打包（PyInstaller --onefile）下，__file__ 指向 _MEIPASS 临时目录，
# 数据/HTML 应放在「exe 自身所在目录」（即 dist/，分发时拷贝过去），故冻结态改用 sys.executable 目录
if getattr(sys, "frozen", False):
    HERE = os.path.dirname(os.path.abspath(sys.executable))
else:
    HERE = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(HERE, 'pending.json')
HTML_FILE = os.path.join(HERE, 'inspector_matcher.html')
STANDARDS_DIR = os.path.join(HERE, 'standards')
STANDARDS_INDEX = os.path.join(STANDARDS_DIR, 'standards_index.json')
PORT = 8721

# 是否为 GUI/托盘模式：无 stdout（pythonw 启动 或 打包 --noconsole）即视为托盘模式，
# 主线程跑托盘图标、无控制台窗口；有 stdout（python.exe 控制台）则前台打印横幅。
IS_GUI_MODE = (sys.stdout is None)


def _redirect_none_streams():
    """pythonw（无控制台）下 sys.stdout/stderr 为 None，任何 print / 日志写入都会抛
    AttributeError 并破坏 HTTP 响应（典型表现：GET / 返回 0 字节 → 浏览器空白）。
    这里在启动期把 None 的流重定向到调试日志，彻底消除该隐患。"""
    if sys.stdout is not None and sys.stderr is not None:
        return
    try:
        _f = open(os.path.join(HERE, 'server_debug.log'), 'a', encoding='utf-8', buffering=1)
    except Exception:
        try:
            _f = open(os.devnull, 'w')
        except Exception:
            return
    if sys.stdout is None:
        sys.stdout = _f
    if sys.stderr is None:
        sys.stderr = _f


_redirect_none_streams()

# 解析 / 模板生成依赖 openpyxl；若环境缺少则标准上传/下载不可用，但核心功能不受影响
try:
    import standard_lib
    _HAS_STDLIB = True
except Exception:
    _HAS_STDLIB = False

# 阶段四：免费 LLM 语义匹配
try:
    import ai_lib
    _HAS_AI = True
except Exception:
    _HAS_AI = False

# 托盘依赖（缺失则退化为只跑 HTTP server，前台模式打印启动信息）
try:
    import pystray
    from PIL import Image, ImageDraw
    _HAS_TRAY = True
except Exception:
    _HAS_TRAY = False

# ---------------- 随机补足（随机问题清单.xlsx） ----------------
import re
from collections import Counter, defaultdict

CHECKLIST_FILE = os.path.join(HERE, '随机问题清单.xlsx')
USER_CHECKLIST_FILE = os.path.join(HERE, '随机问题清单_用户补充.xlsx')
RANDOM_EXCLUDE_FILE = os.path.join(HERE, '随机问题清单_剔除.json')
_checklist_cache = {"key": None, "items": []}

# 一级分类关键词（顺序敏感：靠前的优先匹配；命中即定类，未命中归「其他」）。
# 若 xlsx 第二列自带「一级分类」，则以第二列为准，跳过本规则。
_CAT_RULES = [
    ("厕所及清洁", ["厕所", "卫生间", "清洁", "污渍", "蛛网", "积水", "卫生", "杂物", "杂草", "垃圾", "整洁", "环境"]),
    ("标识及宣传", ["标识", "海报", "宣传", "物料", "张贴", "卷边", "褪色"]),
    ("便利店及商品", ["便利店", "商品", "价签", "临期", "陈列", "货架", "保健食品", "缺货", "丰满", "货品"]),
    ("员工及服务", ["员工", "着装", "工牌", "服务动作", "开口营销", "站经理", "班组", "营销", "客户", "走访", "便民服务", "司机之家", "爱心驿站", "值守"]),
    ("台账及记录", ["台账", "记录", "日志", "盘点", "归档", "票据", "单据", "签字", "经营分析", "考核", "公示", "看板", "宣贯", "自查"]),
    ("安全及消防", ["消防", "应急", "配电", "保险柜", "资金", "监控", "密码", "视频", "静电", "跨接", "接地", "警示牌", "应急灯", "沙", "现金"]),
    ("加油区及设备", ["加油机", "加油区", "泵岛", "管线", "储罐", "卸油", "液位仪", "铅封", "服务箱"]),
    ("其他", []),
]


def _classify(text):
    for cat, kws in _CAT_RULES:
        if cat == "其他":
            return "其他"
        for kw in kws:
            if kw in text:
                return cat
    return "其他"


# 视频督导模式专用过滤词：仅命中这些词的「现场服务类」问题才允许被补足。
# 锚定「六步法 / 五步法」本身，并覆盖其典型服务步骤动作（避免只靠字面二字漏选）。
# 聚焦服务执行步骤（问候/引车/双手/唱收唱付/微笑/引导/推介/自助加油/洗车等），
# 不含泛化的「增值服务」（更接近营销而非六步法/五步法本身）。
# 注意：刻意不含裸「收银」（会误命中「收银台存放现金」这类资金安全问题），
# 室内收银五步法已由「五步法」覆盖。
_VIDEO_KW = [
    "六步法", "五步法", "服务流程", "服务动作", "开口营销", "微笑",
    "问候", "迎客", "送别", "引车", "双手", "唱收", "唱付",
    "引导", "推介", "自助加油", "洗车",
]


def _is_video_service(text):
    """视频督导模式：判断一条问题是否属于「六步法 / 五步法」相关现场服务问题。"""
    t = text or ""
    return any(kw in t for kw in _VIDEO_KW)


def _norm(s):
    s = (s or '').lower()
    s = re.sub(r'[\s，。、；：,.;:!！?？()（）\[\]【】""\'\'\"\-_/\\|]+', '', s)
    return s


def _similar(a, b):
    """用于「补充内容尽量不与手工内容重复」：归一化相等 / 互相包含 / 短文本被长文本包含（字符二元组
    包含度 > 0.7）均视为相似。用「包含度」而非对称 Jaccard，避免「手工短句恰好是清单长句片段」漏判。"""
    na, nb = _norm(a), _norm(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    if na in nb or nb in na:
        return True
    def grams(x):
        return set(x[i:i + 2] for i in range(len(x) - 1)) if len(x) > 1 else set(x)
    A, B = grams(na), grams(nb)
    if not A or not B:
        return False
    # 包含度：交集 / 较短者规模。短句完全落在长句内时 = 1.0，能准确识别片段关系。
    return len(A & B) / min(len(A), len(B)) > 0.7


_CHECKLIST_HEADER = {'随机问题清单', '问题描述', '一级分类', '内容', '分类', '类别', '序号', '问题', '编号'}


def _load_one_checklist(path):
    """读取单个 xlsx，返回 [{"text","cat1"}]。文件不存在或有误返回 []。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        max_col = ws.max_column
        items = []
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            a = str(a).strip() if a is not None else ''
            if not a or a in _CHECKLIST_HEADER:
                continue
            cat = None
            if max_col >= 2:
                b = ws.cell(row=r, column=2).value
                b = str(b).strip() if b is not None else ''
                if b:
                    cat = b
            items.append({"text": a, "cat1": cat or _classify(a)})
        return items
    except Exception as e:
        _write_debug("checklist load error (%s): %s" % (path, e))
        return []


def _load_checklist():
    """合并主清单(随机问题清单.xlsx) + 用户补充(随机问题清单_用户补充.xlsx)。
    任一文件 mtime 变化即刷新缓存。"""
    try:
        m1 = os.path.getmtime(CHECKLIST_FILE)
    except OSError:
        m1 = None
    try:
        m2 = os.path.getmtime(USER_CHECKLIST_FILE)
    except OSError:
        m2 = None
    key = (m1, m2)
    if key == _checklist_cache["key"]:
        return _checklist_cache["items"]
    items = []
    main = _load_one_checklist(CHECKLIST_FILE)
    if main is None:
        _write_debug("checklist missing: %s" % CHECKLIST_FILE)
    else:
        items.extend(main)
    user = _load_one_checklist(USER_CHECKLIST_FILE)
    if user:
        items.extend(user)
    _checklist_cache["key"] = key
    _checklist_cache["items"] = items
    _write_debug("checklist loaded: %d items (main=%d user=%d)" % (
        len(items), len(main) if main else 0, len(user) if user else 0))
    return items


def _load_exclude():
    """已剔除（不希望被随机补足）的条目，以归一化文本列表存储，可逆。"""
    try:
        if os.path.exists(RANDOM_EXCLUDE_FILE):
            with open(RANDOM_EXCLUDE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, list):
                return [str(x) for x in data if x]
    except Exception:
        pass
    return []


def _save_exclude(norms):
    with open(RANDOM_EXCLUDE_FILE, 'w', encoding='utf-8') as f:
        json.dump(norms, f, ensure_ascii=False, indent=2)


def _random_list():
    """返回全部随机清单条目（主+补充），标注来源与是否已剔除，供前端手动管理。"""
    exset = set(_load_exclude())
    items = []
    for src, path in (('main', CHECKLIST_FILE), ('user', USER_CHECKLIST_FILE)):
        for it in _load_one_checklist(path):
            nt = _norm(it["text"])
            items.append({
                "text": it["text"],
                "cat1": it["cat1"],
                "source": src,
                "norm": nt,
                "excluded": nt in exset,
            })
    return {
        "ok": True,
        "items": items,
        "excluded": list(exset),
        "total": len(items),
        "excluded_count": len(exset),
    }


def _parse_upload_text(raw):
    """纯文本 / txt：每行一条；行内以 Tab 分隔第二列可作「一级分类」。"""
    rows = []
    for line in (raw or '').splitlines():
        s = line.strip()
        if not s:
            continue
        if s in ('随机问题清单', '问题描述', '内容', '一级分类', '分类', '类别'):
            continue
        if '\t' in s:
            a, _, b = s.partition('\t')
            a = a.strip(); b = b.strip()
            if a:
                rows.append((a, b or None))
                continue
        rows.append((s, None))
    return rows


def _decode_text(raw):
    """上传文件字节 → 文本。Excel/浏览器导出的 CSV 多为 GBK/GB18030（尤以 Windows 为中文环境），
    直接用 utf-8 解会乱码；这里优先 utf-8-sig，失败（含 BOM/非法序列）则回退 GB18030，最后才 replace。"""
    try:
        return raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        pass
    try:
        return raw.decode('gb18030')
    except UnicodeDecodeError:
        return raw.decode('utf-8', errors='replace')


def _extract_rows(raw, filename):
    """从上传字节解析出 [(text, cat)]，支持 xlsx/xls/csv/txt。
    针对常见巡检清单格式做了鲁棒处理：跳过表头行、丢弃纯数字序号列、
    取最长文本列作为「问题描述」、取较短文本作为「一级分类」。"""
    fname = (filename or '').lower()
    _HEADER_KW = ('序号', '编号', 'no', 'id', '一级分类', '二级分项',
                  '问题描述', '问题', '内容', '分类', '类别')
    if fname.endswith(('.xlsx', '.xls')) or raw[:2] == b'PK':
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        rows = []
        for r in range(1, ws.max_row + 1):
            entries = []  # (col_idx, text)
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                t = str(v).strip()
                if t:
                    entries.append((c, t))
            if not entries:
                continue
            # 跳过表头行（首格是常见表头关键词）
            if entries[0][1] in _HEADER_KW:
                continue
            # 首列是纯数字序号 → 丢弃，内容取后续最长文本列
            if entries[0][1].isdigit():
                entries = entries[1:]
            if not entries:
                continue
            content = max(entries, key=lambda e: len(e[1]))[1]
            cat = None
            for _, t in entries:
                if t != content and len(t) <= 12 and '。' not in t and '（' not in t and not t.isdigit():
                    cat = t
                    break
            rows.append((content, cat))
        return rows
    if fname.endswith('.csv'):
        import csv
        text = _decode_text(raw)
        rows = []
        for rec in csv.reader(io.StringIO(text)):
            if not rec:
                continue
            a = rec[0].strip()
            if not a or a == '随机问题清单':
                continue
            if a in _HEADER_KW:
                continue
            if len(rec) >= 3:
                # 3 列格式：一级分类, 问题描述, 明细（明细常为编号多行，拆成多条）
                cat = rec[0].strip()
                prob = rec[1].strip()
                detail = rec[2].strip()
                if not prob and not detail:
                    continue
                # 明细按换行拆成独立小问题（去掉 "1. " 序号），各自成一条
                subs = [re.sub(r'^\d+[\.、]\s*', '', ln.strip()).strip()
                        for ln in detail.split('\n')]
                subs = [s for s in subs if s]
                if subs:
                    for s in subs:
                        rows.append((s, cat or None))
                    continue
                if prob:
                    rows.append((prob, cat or None))
            else:
                cat = rec[1].strip() if len(rec) > 1 else ''
                rows.append((a, cat or None))
        return rows
    return _parse_upload_text(_decode_text(raw))


def _parse_multipart(body, boundary):
    """极简 multipart/form-data 解析，返回 {name: (filename, content_bytes)}。"""
    parts = {}
    delimiter = b'--' + boundary
    for seg in body.split(delimiter):
        if seg in (b'', b'--', b'\r\n'):
            continue
        if seg.startswith(b'\r\n'):
            seg = seg[2:]
        if seg.endswith(b'\r\n'):
            seg = seg[:-2]
        if b'\r\n\r\n' not in seg:
            continue
        head, content = seg.split(b'\r\n\r\n', 1)
        name = None
        filename = None
        for line in head.split(b'\r\n'):
            ls = line.decode('utf-8', 'replace').lower()
            if 'content-disposition' in ls:
                s = line.decode('utf-8', 'replace')
                m = re.search(r'name="([^"]*)"', s)
                if m:
                    name = m.group(1)
                m = re.search(r'filename="([^"]*)"', s)
                if m:
                    filename = m.group(1)
        if name is not None:
            parts[name] = (filename, content)
    return parts


def _select_fill(pool, need, max_per_cat):
    """从候选池按「一级分类」分散选取，同一分类最多 max_per_cat 条（减少同类重复）。"""
    import random
    if need <= 0:
        return []
    by_cat = defaultdict(list)
    for it in pool:
        by_cat[it["cat1"]].append(it)
    for c in by_cat:
        random.shuffle(by_cat[c])
    selected = []
    used = Counter()
    remaining = need
    while remaining > 0:
        avail = [c for c in by_cat if by_cat[c] and used[c] < max_per_cat]
        if not avail:
            break
        # 优先选「当前被选次数最少」的分类，平局随机 → 自然分散到不同一级分类
        avail.sort(key=lambda c: (used[c], random.random()))
        c = avail[0]
        selected.append(by_cat[c].pop(0))
        used[c] += 1
        remaining -= 1
    # 最终再整体洗牌一次：避免补足结果按「分类」成簇排列，让每次补足的呈现顺序也完全随机
    random.shuffle(selected)
    return selected


def _random_fill(manual, target, max_per_cat, video_mode=False):
    manual = [str(x).strip() for x in (manual or []) if str(x).strip()]
    # 视频督导模式：随机补足只补 2 条，且只取「六步法/五步法」相关现场服务问题
    if video_mode:
        target = 2
        max_per_cat = 2
    need = max(0, int(target) - len(manual))
    items = _load_checklist()
    excluded = set(_load_exclude())
    # 候选池：剔除与手工内容相似的；剔除被手动屏蔽的；池内按归一化文本去重
    pool = []
    seen = set()
    for it in items:
        t = it["text"]
        if any(_similar(t, m) for m in manual):
            continue
        nt = _norm(t)
        if nt in seen:
            continue
        if nt in excluded:
            continue
        # 视频督导模式：仅保留「六步法/五步法」相关现场服务问题
        if video_mode and not _is_video_service(t):
            continue
        seen.add(nt)
        pool.append(it)
    filled = _select_fill(pool, need, max_per_cat)
    if video_mode:
        if filled:
            note = "🎥 视频督导模式：已从《随机问题清单》补足 %d 条「六步法/五步法」相关现场服务问题（与手工内容去重）。" % len(filled)
        elif need <= 0:
            note = "手工已≥%d 条，无需补足。" % target
        elif not pool:
            note = "🎥 视频督导模式：清单中无「六步法/五步法」相关现场服务问题可补足（或均被剔除/与手工重复）。"
        else:
            note = "🎥 视频督导模式：可用条目已用尽或均与手工重复，无法补足到 %d 条。" % target
    else:
        if filled:
            note = "已从《随机问题清单》随机补足 %d 条（一级分类单类≤%d，且与手工内容去重）。" % (len(filled), max_per_cat)
        elif need <= 0:
            note = "手工已≥%d 条，无需补足。" % target
        else:
            note = "清单可用条目已用尽或均与手工重复，无法补足到 %d 条。" % target
    return {
        "ok": True,
        "manual": len(manual),
        "need": need,
        "filled": [{"text": f["text"], "cat1": f["cat1"]} for f in filled],
        "total": len(manual) + len(filled),
        "video_mode": video_mode,
        "note": note,
    }


_lock = threading.Lock()
_server = None  # 持有 ThreadingHTTPServer 引用，托盘"退出"时可调 shutdown()


def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {"items": [], "version": 1, "updated_at": None}


def save_data(data):
    data['updated_at'] = int(time.time() * 1000)
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ---------------- 检查标准存储 ----------------
def ensure_standards_dir():
    os.makedirs(STANDARDS_DIR, exist_ok=True)
    if not os.path.exists(STANDARDS_INDEX):
        with open(STANDARDS_INDEX, 'w', encoding='utf-8') as f:
            json.dump([], f, ensure_ascii=False)


def load_standards_index():
    ensure_standards_dir()
    try:
        with open(STANDARDS_INDEX, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []


def save_standards_index(index):
    ensure_standards_dir()
    with open(STANDARDS_INDEX, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=2)


def _std_path(std_id):
    return os.path.join(STANDARDS_DIR, std_id + '.json')


class Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        # pythonw 下 sys.stderr 为 None，改为写入调试日志（带 self 检查避免再次抛错）
        try:
            _write_debug("[req][%s] %s" % (self.address_string(), fmt % args))
        except Exception:
            pass

    def _set_cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS, DELETE')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, X-Requested-With')
        self.send_header('Access-Control-Max-Age', '86400')

    def _send_json(self, obj, code=200):
        body = json.dumps(obj, ensure_ascii=False).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self._set_cors()
        self.end_headers()
        self.wfile.write(body)

    def _send_bytes(self, body, content_type, code=200):
        if isinstance(body, str):
            body = body.encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Length', str(len(body)))
        self._set_cors()
        self.end_headers()
        self.wfile.write(body)

    def _read_json(self):
        length = int(self.headers.get('Content-Length', 0))
        if length == 0:
            return {}
        return json.loads(self.rfile.read(length).decode('utf-8'))

    def do_OPTIONS(self):
        self.send_response(204)
        self._set_cors()
        self.end_headers()

    # ---------------- 检查标准 API ----------------
    def _standards_get_list(self):
        index = load_standards_index()
        self._send_json({"ok": True, "standards": index})

    def _standards_get_one(self, std_id):
        path = _std_path(std_id)
        if not os.path.isfile(path):
            self._send_json({"ok": False, "error": "not found"}, 404)
            return
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        self._send_json({"ok": True, "id": std_id, **data})

    def _standards_get_one_dict(self, std_id):
        """仅读取并返回标准 dict（不发送响应），供导出等复用。"""
        path = _std_path(std_id)
        if not os.path.isfile(path):
            return None
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def _standards_upload(self):
        if not _HAS_STDLIB:
            self._send_json({"ok": False, "error": "缺少 openpyxl，无法解析 xlsx"}, 500)
            return
        payload = self._read_json()
        name = (payload.get('name') or '').strip() or '未命名检查标准'
        data_b64 = payload.get('data') or ''
        try:
            raw = base64.b64decode(data_b64)
        except Exception:
            self._send_json({"ok": False, "error": "文件数据解码失败"}, 400)
            return
        try:
            items, cat1_list, cat2_list = standard_lib.parse_xlsx_to_items(io.BytesIO(raw))
        except Exception as e:
            self._send_json({"ok": False, "error": "解析失败：" + str(e)}, 400)
            return
        if not items:
            self._send_json({"ok": False, "error": "未解析到任何条目，请确认使用了正确的导入模板"}, 400)
            return
        std_id = 'std_' + uuid.uuid4().hex[:10]
        rec = {
            "id": std_id, "name": name,
            "n": len(items), "cat1_list": cat1_list, "cat2_list": cat2_list,
            "items": items, "created_at": int(time.time() * 1000),
        }
        with open(_std_path(std_id), 'w', encoding='utf-8') as f:
            json.dump(rec, f, ensure_ascii=False, indent=1)
        index = load_standards_index()
        index.append({"id": std_id, "name": name, "n": len(items)})
        save_standards_index(index)
        self._send_json({"ok": True, "id": std_id, "name": name, "n": len(items)})

    def _standards_delete(self, std_id):
        path = _std_path(std_id)
        if os.path.isfile(path):
            os.remove(path)
        index = [s for s in load_standards_index() if s.get('id') != std_id]
        save_standards_index(index)
        self._send_json({"ok": True, "removed": std_id})

    def _standards_template(self):
        if not _HAS_STDLIB:
            self._send_json({"ok": False, "error": "缺少 openpyxl"}, 500)
            return
        from urllib.parse import quote
        # 用内置标准(items.json)填充示例，避免空白模板让人误以为下载失败
        items = []
        try:
            with open(os.path.join(HERE, "items.json"), encoding="utf-8") as f:
                items = json.load(f).get("items", []) or []
        except Exception:
            pass
        body = standard_lib.make_template_bytes(
            items,
            title="检查标准条目导入模板（下表为当前内置标准示例，可参照修改；"
                  "也可整列清空后填写你自己的标准。导入时后台会自动识别列并转换，"
                  "无论表头在第几行都能识别。）",
        )
        fname = '检查标准导入模板.xlsx'
        disp = 'attachment; filename="inspection_standard_template.xlsx"; filename*=UTF-8\'\'%s' % quote(fname)
        self.send_response(200)
        self.send_header('Content-Type',
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', disp)
        self.send_header('Content-Length', str(len(body)))
        self._set_cors()
        self.end_headers()
        self.wfile.write(body)

    def _standards_export(self, std_id):
        """把一个标准（内置或已上传）导出为填好内容的 xlsx，即「自动转换」后的规范文件。"""
        if not _HAS_STDLIB:
            self._send_json({"ok": False, "error": "缺少 openpyxl"}, 500)
            return
        from urllib.parse import quote
        if std_id == "builtin":
            try:
                with open(os.path.join(HERE, "items.json"), encoding="utf-8") as f:
                    data = json.load(f)
            except Exception as e:
                self._send_json({"ok": False, "error": str(e)}, 500)
                return
            items = data.get("items", []) or []
            fname = '内置检查标准.xlsx'
        else:
            rec = self._standards_get_one_dict(std_id)
            if not rec:
                self._send_json({"ok": False, "error": "标准不存在"}, 404)
                return
            items = rec.get("items", []) or []
            fname = (rec.get("name") or "检查标准") + ".xlsx"
        body = standard_lib.make_template_bytes(
            items, title="检查标准导出（后台自动转换后的规范格式，可直接再次导入）")
        disp = 'attachment; filename="%s"; filename*=UTF-8\'\'%s' % (
            quote(fname), quote(fname))
        self.send_response(200)
        self.send_header('Content-Type',
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', disp)
        self.send_header('Content-Length', str(len(body)))
        self._set_cors()
        self.end_headers()
        self.wfile.write(body)

    # ---------------- 随机清单扩充上传 ----------------
    def _random_upload(self):
        """扩充随机问题清单：支持文件上传(xlsx/csv/txt)或文本粘贴，append/replace 模式。
        append -> 写入/合并到 随机问题清单_用户补充.xlsx；replace -> 仅保留本次内容。
        主清单(随机问题清单.xlsx)始终不动。"""
        ctype = self.headers.get('Content-Type', '')
        mode = 'append'
        new_rows = []
        src_desc = ''
        if 'multipart/form-data' in ctype:
            m = re.search(r'boundary=([^;]+)', ctype)
            if not m:
                self._send_json({"ok": False, "error": "multipart 缺少 boundary"}, 400)
                return
            boundary = m.group(1).strip().encode('utf-8')
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length) if length else b''
            parts = _parse_multipart(body, boundary)
            mode = (parts.get('mode') or (None, b'append'))[1].decode('utf-8', 'replace').strip() or 'append'
            if 'file' in parts and parts['file'][1]:
                fname, raw = parts['file']
                try:
                    new_rows = _extract_rows(raw, fname or 'upload')
                except Exception as e:
                    self._send_json({"ok": False, "error": "文件解析失败：" + str(e)}, 400)
                    return
                src_desc = fname or '文件'
            elif 'text' in parts and parts['text'][1].strip():
                new_rows = _parse_upload_text(parts['text'][1].decode('utf-8', 'replace'))
                src_desc = '文本粘贴'
        else:
            payload = self._read_json()
            mode = (payload.get('mode') or 'append')
            if payload.get('rows'):
                for row in payload['rows']:
                    if isinstance(row, list) and row:
                        t = str(row[0]).strip()
                        c = str(row[1]).strip() if len(row) > 1 else ''
                        if t:
                            new_rows.append((t, c or None))
            elif payload.get('text'):
                new_rows = _parse_upload_text(payload['text'])
            src_desc = '文本粘贴'
        if not new_rows:
            self._send_json({"ok": False, "error": "没有解析到任何条目"}, 400)
            return
        # 与现有清单去重（相似即跳过）
        seen = set(_norm(it["text"]) for it in _load_checklist())
        merged = []
        added = 0
        for t, c in new_rows:
            nt = _norm(t)
            if not nt or nt in seen:
                continue
            seen.add(nt)
            merged.append((t, c or _classify(t)))
            added += 1
        if added == 0:
            self._send_json({"ok": True, "added": 0, "skipped": len(new_rows),
                             "message": "提交的条目都已存在于清单中，未新增。"})
            return
        # 组装写入内容
        if mode == 'replace':
            base = merged
        else:
            # append：保留现有补充全部条目，仅跳过与「本次新增项」完全相同的（避免重复）。
            # 注意：不能用 seen（它已含全部现有清单）来过滤现有项，否则会把原有条目全部丢弃，
            # 导致 append 退化成 replace（曾有数据丢失 bug）。
            merged_norms = set(_norm(m[0]) for m in merged)
            base = []
            u = _load_one_checklist(USER_CHECKLIST_FILE)
            if u:
                for it in u:
                    if _norm(it["text"]) in merged_norms:
                        continue
                    base.append((it["text"], it["cat1"]))
            base.extend(merged)
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            for t, c in base:
                ws.append([t, c])
            wb.save(USER_CHECKLIST_FILE)
        except Exception as e:
            self._send_json({"ok": False, "error": "写入补充清单失败：" + str(e)}, 500)
            return
        # 立即失效缓存，保证下次随机补足读到新内容
        global _checklist_cache
        _checklist_cache = {"key": None, "items": []}
        total_all = len(_load_checklist())
        self._send_json({
            "ok": True,
            "added": added,
            "skipped": len(new_rows) - added,
            "total_user": len(base),
            "total_all": total_all,
            "message": "已%s %d 条（跳过 %d 条重复）；用户补充共 %d 条，全清单共 %d 条。" % (
                '替换' if mode == 'replace' else '追加', added, len(new_rows) - added,
                len(base), total_all),
        })

    # ---------------- 阶段四：AI 语义匹配 ----------------
    def _ai_config_get(self):
        if not _HAS_AI:
            self._send_json({"ok": False, "error": "AI 模块加载失败"}, 500)
            return
        cfg = ai_lib.load_config()
        # 不回显密钥：只返回是否已配置，前端据此决定是否提示填写
        self._send_json({
            "ok": True,
            "provider": cfg.get("provider"),
            "enabled": cfg.get("enabled"),
            "base_url": cfg.get("base_url"),
            "model": cfg.get("model"),
            "has_key": bool(cfg.get("api_key")),
            "configured": ai_lib.is_configured(cfg),
        })

    def _ai_config_post(self):
        if not _HAS_AI:
            self._send_json({"ok": False, "error": "AI 模块加载失败"}, 500)
            return
        payload = self._read_json()
        cfg = ai_lib.load_config()
        provider = payload.get("provider") or cfg.get("provider") or "deepseek"
        cfg = ai_lib.apply_preset(cfg, provider)

        if payload.get("base_url"):
            cfg["base_url"] = payload["base_url"].strip()
        if payload.get("model"):
            cfg["model"] = payload["model"].strip()
        if "enabled" in payload:
            cfg["enabled"] = bool(payload["enabled"])

        # api_key 处理：传 null/不传 => 保留原值；传非空 => 更新；传 "__CLEAR__" => 清空
        ak = payload.get("api_key")
        if ak is None:
            pass
        elif ak == "__CLEAR__":
            cfg["api_key"] = ""
        elif ak != "":
            cfg["api_key"] = ak

        ai_lib.save_config(cfg)
        cfg = ai_lib.load_config()
        self._send_json({
            "ok": True,
            "provider": cfg.get("provider"),
            "enabled": cfg.get("enabled"),
            "configured": ai_lib.is_configured(cfg),
            "has_key": bool(cfg.get("api_key")),
        })

    def _ai_match_post(self):
        if not _HAS_AI:
            self._send_json({"ok": False, "error": "AI 模块加载失败"}, 500)
            return
        payload = self._read_json()
        records = payload.get("records") or []
        if not isinstance(records, list) or not records:
            self._send_json({"ok": False, "error": "records 为空"}, 400)
            return
        cfg = ai_lib.load_config()
        if not cfg.get("enabled") and cfg.get("provider") != "mock":
            self._send_json({"ok": False, "error": "AI 未启用，请在 ⚙️ AI 设置 中开启"}, 400)
            return
        try:
            results = ai_lib.match_records(cfg, records)
        except Exception as e:
            self._send_json({"ok": False, "error": "AI 匹配失败：" + str(e)}, 500)
            return
        self._send_json({"ok": True, "results": results})

    # ---------------- 路由 ----------------
    def do_GET(self):
        url = urlparse(self.path)
        path = url.path
        if path == '/api/pending':
            self._send_json(load_data())
        elif path == '/api/health':
            self._send_json({"ok": True, "ts": int(time.time())})
        elif path == '/api/standards':
            self._standards_get_list()
        elif path == '/api/standards/template':
            self._standards_template()
        elif path == '/api/random_list':
            self._send_json(_random_list())
            return
        elif path == '/api/ai/config':
            self._ai_config_get()
        elif path.endswith('/xlsx') and path.startswith('/api/standards/'):
            # /api/standards/<id>/xlsx  -> 导出该标准（自动转换后的规范 xlsx）
            std_id = path[len('/api/standards/'):-len('/xlsx')].strip('/')
            if std_id:
                self._standards_export(std_id)
            else:
                self._send_json({"ok": False, "error": "bad id"}, 400)
        elif path.startswith('/api/standards/'):
            std_id = path[len('/api/standards/'):].strip('/')
            if std_id:
                self._standards_get_one(std_id)
            else:
                self._send_json({"ok": False, "error": "bad id"}, 400)
        elif path in ('/', '/index.html'):
            if os.path.exists(HTML_FILE):
                with open(HTML_FILE, 'rb') as f:
                    self._send_bytes(f.read(), 'text/html; charset=utf-8')
            else:
                self.send_error(404, 'inspector_matcher.html not found')
        elif path.startswith('/static/'):
            rel = path[len('/static/'):]
            full = os.path.normpath(os.path.join(HERE, rel))
            if not full.startswith(HERE):
                self.send_error(403)
                return
            if os.path.isfile(full):
                if full.endswith('.js'):
                    ct = 'application/javascript; charset=utf-8'
                elif full.endswith('.css'):
                    ct = 'text/css; charset=utf-8'
                elif full.endswith('.json'):
                    ct = 'application/json; charset=utf-8'
                else:
                    ct = 'text/plain; charset=utf-8'
                with open(full, 'rb') as f:
                    self._send_bytes(f.read(), ct)
            else:
                self.send_error(404)
        else:
            self.send_error(404)

    def do_POST(self):
        url = urlparse(self.path)
        path = url.path
        if path == '/api/standards/upload':
            self._standards_upload()
            return
        if path == '/api/ai/config':
            self._ai_config_post()
            return
        if path == '/api/ai/match':
            self._ai_match_post()
            return
        if path == '/api/random_fill':
            payload = self._read_json()
            self._send_json(_random_fill(
                payload.get('manual'),
                payload.get('target', 10),
                payload.get('max_per_cat', 3),
                payload.get('video_mode', False),
            ))
            return
        if path == '/api/random_upload':
            self._random_upload()
            return
        if path == '/api/random_exclude':
            payload = self._read_json()
            norms = []
            for t in (payload.get('texts') or []):
                nt = _norm(t)
                if nt:
                    norms.append(nt)
            norms = list(dict.fromkeys(norms))  # 去重保序
            _save_exclude(norms)
            global _checklist_cache
            _checklist_cache = {"key": None, "items": []}
            self._send_json({
                "ok": True,
                "excluded": len(norms),
                "message": "已更新剔除名单，共 %d 条；下次随机补足即生效。" % len(norms),
            })
            return
        if path == '/api/shutdown':
            try:
                if _server is not None:
                    _server.shutdown()
            except Exception:
                pass
            self._send_json({"ok": True})

            def _delay_exit():
                time.sleep(0.3)
                os._exit(0)

            threading.Thread(target=_delay_exit, daemon=True).start()
            return
        with _lock:
            data = load_data()
            if path == '/api/push':
                payload = self._read_json()
                items = payload.get('items', [])
                if not isinstance(items, list):
                    self._send_json({"ok": False, "error": "items must be list"}, 400)
                    return
                by_code = {it.get('code'): it for it in data['items'] if it.get('code')}
                added = updated = 0
                for it in items:
                    if not isinstance(it, dict):
                        continue
                    if 'id' not in it:
                        it['id'] = it.get('code') or str(uuid.uuid4())[:8]
                    it['added_at'] = int(time.time() * 1000)
                    code = it.get('code')
                    if code and code in by_code:
                        by_code[code].update(it)
                        updated += 1
                    else:
                        data['items'].append(it)
                        if code:
                            by_code[code] = it
                        added += 1
                save_data(data)
                self._send_json({"ok": True, "n": len(data['items']), "added": added, "updated": updated})
            elif path == '/api/remove':
                payload = self._read_json()
                item_id = payload.get('id')
                before = len(data['items'])
                data['items'] = [it for it in data['items'] if it.get('id') != item_id]
                save_data(data)
                self._send_json({"ok": True, "n": len(data['items']), "removed": before - len(data['items'])})
            elif path == '/api/clear':
                data['items'] = []
                save_data(data)
                self._send_json({"ok": True, "n": 0})
            else:
                self.send_error(404)

    def do_DELETE(self):
        url = urlparse(self.path)
        path = url.path
        if path.startswith('/api/standards/'):
            std_id = path[len('/api/standards/'):].strip('/')
            if std_id:
                self._standards_delete(std_id)
                return
        self.send_error(404)


def _make_tray_icon():
    """生成 64x64 托盘图标（程序内绘制，避免依赖外部 PNG）"""
    img = Image.new('RGBA', (64, 64), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)
    # 蓝色圆角底
    d.rounded_rectangle((4, 4, 60, 60), radius=12, fill=(37, 99, 235, 255))
    # 白色"表"字简笔
    d.rectangle((18, 14, 46, 18), fill=(255, 255, 255, 255))   # 横
    d.rectangle((18, 30, 46, 34), fill=(255, 255, 255, 255))   # 横
    d.rectangle((18, 46, 46, 50), fill=(255, 255, 255, 255))   # 横
    d.rectangle((30, 14, 34, 50), fill=(255, 255, 255, 255))   # 竖
    return img


def _open_browser():
    """确认本服务真正在监听后，再打开浏览器，避免「浏览器打开时端口还没就绪」导致白屏。
    注意：urllib 默认会读取系统/IE 代理设置，localhost 常被误路由到代理而连不上，
    这里显式禁用代理、强制直连 127.0.0.1，否则会误判服务未就绪（日志里 ready=False）。"""
    import urllib.request
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    deadline = time.time() + 8.0
    ready = False
    while time.time() < deadline:
        try:
            with opener.open("http://127.0.0.1:%d/api/health" % PORT, timeout=0.5) as r:
                if r.status == 200:
                    ready = True
                    break
        except Exception:
            pass
        time.sleep(0.3)
    _write_debug("browser open: server ready=%s on port=%d" % (ready, PORT))
    try:
        webbrowser.open("http://127.0.0.1:%d" % PORT)
    except Exception:
        # 没有默认浏览器时也给出提示，而不是静默白屏
        _err_box(u"无法自动打开浏览器，请手动在浏览器访问：\nhttp://127.0.0.1:%d" % PORT)


def _run_tray():
    """在主线程跑托盘图标。菜单：打开网页 / 退出。
    HTTP server 已在子线程启动，这里只是阻塞等待托盘消息循环。"""
    icon = pystray.Icon(
        name="inspector_matcher",
        icon=_make_tray_icon(),
        title="督导填表助手 - 已运行（http://127.0.0.1:%d）" % PORT,
        menu=pystray.Menu(
            pystray.MenuItem("🌐 打开网页", lambda: webbrowser.open("http://127.0.0.1:%d" % PORT)),
            pystray.MenuItem("📂 打开数据目录", lambda: _open_data_dir()),
            pystray.Menu.SEPARATOR,
            pystray.MenuItem("❌ 退出服务", lambda: _quit_app(icon)),
        ),
    )
    # 双击托盘图标 = 打开网页
    icon.default_action = lambda: webbrowser.open("http://127.0.0.1:%d" % PORT)
    icon.run()


def _open_data_dir():
    """打开数据所在目录（Windows 资源管理器；其它系统用 xdg-open）"""
    target = HERE
    try:
        if os.name == 'nt':
            os.startfile(target)  # noqa
        elif sys.platform == 'darwin':
            os.system('open "%s"' % target)
        else:
            os.system('xdg-open "%s"' % target)
    except Exception:
        pass


def _quit_app(icon):
    """托盘菜单「退出」：停托盘、停 HTTP server"""
    try:
        if _server is not None:
            _server.shutdown()
    except Exception:
        pass
    try:
        icon.stop()
    except Exception:
        pass
    # server.serve_forever() 返回后 main() 自然结束
    os._exit(0)


def _write_debug(msg):
    """把关键启动/运行信息写入 exe 同目录的 server_debug.log，便于无控制台模式下排错。"""
    try:
        with open(os.path.join(HERE, 'server_debug.log'), 'a', encoding='utf-8') as f:
            f.write('[%s] %s\n' % (time.strftime('%Y-%m-%d %H:%M:%S'), msg))
    except Exception:
        pass


def _pick_port(start=8721, count=10):
    """优先用 start，被占用则顺延到 start+count-1，返回第一个可用端口。"""
    import socket
    for p in range(start, start + count):
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        try:
            s.bind(('127.0.0.1', p))
            return p
        except OSError:
            pass
        finally:
            try:
                s.close()
            except Exception:
                pass
    return None


def _probe_health(url, timeout=1.0):
    """探测某地址是否为「我们自己仍在运行的服务」（/api/health 返回 200）。"""
    try:
        import urllib.request
        with urllib.request.urlopen(url, timeout=timeout) as r:
            return r.status == 200
    except Exception:
        return False


_LOCK_PORT = 8720  # 互斥锁端口：持有即代表唯一实例在跑，保证服务永远固定 8721，避免端口漂移


def _acquire_instance_lock():
    """用 8720 作为跨进程互斥锁，保证全机只有一个 server 实例在跑。
    成功：返回该 socket（保持打开以持有锁）；失败（已被占用）：返回 None。"""
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        s.bind(('127.0.0.1', _LOCK_PORT))
        return s
    except OSError:
        try:
            s.close()
        except Exception:
            pass
        return None


def _write_instance_port(port):
    """把「当前实例实际监听的端口」写到 exe 同目录，供后续双击的实例读取，
    从而打开正确的地址（端口可能因 8721 被占用而顺延到 8722 等）。"""
    try:
        with open(os.path.join(HERE, 'instance_port.txt'), 'w', encoding='utf-8') as f:
            f.write(str(port))
    except Exception:
        pass


def _read_instance_port():
    try:
        with open(os.path.join(HERE, 'instance_port.txt'), 'r', encoding='utf-8') as f:
            return int((f.read() or '').strip())
    except Exception:
        return None



def _info_box(text, title=u"督导填表助手"):
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, text, title, 0x40)
    except Exception:
        pass


def _err_box(text, title=u"督导填表助手 - 错误"):
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, text, title, 0x30)
    except Exception:
        pass


def _tray_worker():
    """在后台线程跑托盘（best-effort）；主线程继续跑 HTTP server，托盘失败不影响服务。"""
    try:
        _run_tray()
    except Exception as e:
        _write_debug("tray exited: %s" % e)


def _safe_print(*args):
    """Windows GBK 控制台下中文可能 UnicodeEncodeError，做降级处理。"""
    line = ' '.join(str(a) for a in args)
    try:
        print(line)
    except UnicodeEncodeError:
        try:
            print(line.encode('ascii', 'replace').decode('ascii'))
        except Exception:
            pass


def _print_banner():
    """前台控制台模式才打印横幅（GUI/托盘模式无 stdout）"""
    if IS_GUI_MODE:
        return
    _safe_print('=' * 64)
    _safe_print('  督导系统自动填表助手 - 本地服务')
    _safe_print('  监听地址: http://127.0.0.1:%d' % PORT)
    _safe_print('  页面文件: %s' % HTML_FILE)
    if os.path.exists(HTML_FILE):
        _safe_print('  页面大小: %d 字节（存在，正常）' % os.path.getsize(HTML_FILE))
    else:
        _safe_print('  [警告] 页面文件未找到！GET / 将返回 404，浏览器会白屏/报错。')
    _safe_print('  数据文件: %s' % DATA_FILE)
    _safe_print('  检查标准: %s' % STANDARDS_DIR)
    _safe_print('=' * 64)
    _safe_print('  油猴脚本需要先在 360极速X 上安装 Tampermonkey：')
    _safe_print('  https://chromewebstore.google.com/detail/tampermonkey/')
    _safe_print('  然后把 ddgl_autofill.user.js 拖入浏览器即可安装')
    _safe_print('  也可在工具页右上角点击「下载油猴脚本」按钮一键下载')
    _safe_print('=' * 64)
    _safe_print('  按 Ctrl+C 停止服务')
    _safe_print('=' * 64)


def _safe_serve_forever():
    """在【后台线程】运行 HTTP 服务；被 shutdown() 唤醒后返回。"""
    _write_debug("serving forever on %d" % PORT)
    try:
        _server.serve_forever()
    except Exception as e:
        _write_debug("serve_forever error: %s" % e)
        _err_box(u"服务运行出错：%s" % e)
    _write_debug("server stopped")


def main():
    global _server, PORT
    ensure_standards_dir()
    _write_debug("DIAG frozen=%s exe=%s HERE=%s html_exists=%s" % (
        getattr(sys, 'frozen', False), sys.executable, HERE, os.path.exists(HTML_FILE)))

    # 单实例：用锁端口 8720 保证全机只有一个「我们的」实例在跑，
    # 避免多次双击互相抢端口导致地址混乱/白屏。
    lock_sock = _acquire_instance_lock()
    if lock_sock is None:
        # 已有实例在跑：读取它实际监听的端口（可能顺延过），打开正确地址后退出。
        port = _read_instance_port() or 8721
        _write_debug("instance lock (port %d) already held -> open existing page on %d and exit" % (_LOCK_PORT, port))
        try:
            webbrowser.open("http://127.0.0.1:%d" % port)
        except Exception:
            pass
        sys.exit(0)

    # 服务端口：优先 8721；若被别的程序（含未退出的旧版 server.exe）占用，则顺延到下一个空闲端口。
    # 因为锁已保证全局只有一个我们的实例，这里顺延不会造成多实例冲突，且浏览器会打开到正确端口。
    chosen = _pick_port(8721, 11)
    if chosen is None:
        _err_box(u"无法在 8721~8731 找到可用端口，请关闭占用该端口的程序后重试。")
        _write_debug("no free port in 8721..8731")
        sys.exit(1)
    PORT = chosen
    _write_instance_port(PORT)
    _print_banner()
    _write_debug("selected port=%d (holding instance lock on %d)" % (PORT, _LOCK_PORT))
    try:
        _server = http.server.ThreadingHTTPServer(('127.0.0.1', PORT), Handler)
    except Exception as e:
        _err_box(u"服务启动失败（端口 %d 被占用）：%s\n\n请先在任务管理器结束所有 server.exe 进程，再重新双击本程序。" % (PORT, e))
        _write_debug("bind failed: %s" % e)
        sys.exit(1)

    # ★ 关键修复：先把 HTTP 服务跑在【后台线程】，确保端口真正在监听，
    #   再去做任何可能阻塞主线程的事（托盘图标、端口漂移提示框、打开浏览器）。
    #   旧逻辑里 serve_forever 在 main 线程且位于 _info_box 之后：一旦端口顺延
    #   （8721 被占用→改用 8722）弹出的提示框是模态对话框，会阻塞 main 线程，
    #   导致 serve_forever 一直不启动，浏览器打开时端口上根本没有服务 → 白屏。
    serving_thread = threading.Thread(target=_safe_serve_forever, daemon=True)
    serving_thread.start()

    # 自动打开浏览器（服务已在后台线程监听，健康检查会立即通过）
    threading.Thread(target=_open_browser, daemon=True).start()

    NO_TRAY = os.environ.get('INSPECTOR_NO_TRAY') == '1'

    if IS_GUI_MODE and _HAS_TRAY and not NO_TRAY:
        # 托盘在主线程跑（消息循环阻塞主线程，正好充当 keep-alive）。
        # 服务已在后台线程运行，托盘初始化失败也不会影响服务。
        if PORT != 8721:
            # 端口顺延时告知用户（服务已在后台线程运行，弹窗不再阻塞服务）
            _info_box(u"本地服务已启动在 http://127.0.0.1:%d\n（默认 8721 被占用，已自动改用此端口）" % PORT)
        try:
            _run_tray()
        except Exception as e:
            _write_debug("tray failed, fall back to idle loop: %s" % e)
            while True:
                time.sleep(3600)
    elif IS_GUI_MODE:
        # 无托盘的 GUI 模式：提示一次（服务已在后台线程运行，弹窗不再阻塞服务）
        _info_box(u"本地服务已启动：\nhttp://127.0.0.1:%d\n\n停止服务：在任务管理器结束 server.exe，或在网页内点「退出服务」。" % PORT)
        while True:
            time.sleep(3600)
    else:
        # 控制台模式：主线程 join 服务线程，保留 Ctrl+C 停止能力
        try:
            serving_thread.join()
        except KeyboardInterrupt:
            _server.shutdown()


if __name__ == '__main__':
    main()
