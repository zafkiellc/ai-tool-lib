# -*- coding: utf-8 -*-
"""
standard_lib.py
---------------
检查标准（条目模板）的通用解析与模板生成，供 server.py 的上传 / 下载接口复用。

- parse_xlsx_to_items(src): 读取一个 xlsx（路径或 BytesIO），抽取条目列表
  采用"表头探测"而非固定行号，兼容官方下发模板与本工具生成的空白模板。
- make_template_bytes(): 生成一份空白的「检查标准导入模板.xlsx」供用户下载填写。
"""
import io
import re

import openpyxl

# 只保留中文 / 英文 / 数字；标点和空白忽略（与前端 charNgrams 保持一致）
RE_KEEP = re.compile(r"[\u4e00-\u9fa5A-Za-z0-9]")

# 标准列名 -> 字段名
FIELD_MAP = {
    "序号": "idx",
    "条目编码": "code",
    "督导体系": "system",
    "一级分类": "cat1",
    "二级分项": "cat2",
    "检查区域": "area",
    "检查单元": "unit",
    "条目名称": "name",
    "问题描述": "desc",
    "检查方法": "method",
    "标准分数下限": "score_lo",
    "标准分数上限": "score_hi",
    "问题重要级别": "severity",
}
# 字段在 xlsx 中的列序号（兜底，找不到表头时使用）
FIELD_COL_FALLBACK = {
    "idx": 1, "code": 2, "system": 3, "cat1": 4, "cat2": 5,
    "area": 6, "unit": 7, "name": 8, "desc": 9, "method": 10,
    "score_lo": 11, "score_hi": 12, "severity": 13,
}


def char_ngrams(text):
    text = "".join(RE_KEEP.findall(text or ""))
    grams = set()
    for ch in text:
        grams.add(ch)
    for i in range(len(text) - 1):
        grams.add(text[i:i + 2])
    return sorted(grams)


def split_title_desc(name):
    """模板里 name = 短标题 + 详细描述（用第一个"。"分隔）。"""
    if not name:
        return "", ""
    m = re.search(r"[）\)]\s*([^。]+?。)", name)
    if m:
        short = name[: m.end()]
        rest = name[m.end():].strip()
        return short, rest
    parts = name.split("。", 1)
    return (parts[0] + "。" if parts else ""), (parts[1].strip() if len(parts) > 1 else "")


def parse_xlsx_to_items(src):
    """读取 xlsx（路径字符串或 BytesIO），返回 (items, cat1_list, cat2_list)。

    items 中不含 grams（前端会按需计算），字段与内置数据保持一致。
    """
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 取第一个工作表

    # 1) 探测表头行：在前 12 行内找同时含「条目名称」与「条目编码/一级分类」的行
    header_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, min(ws.max_column, 24) + 1))
        if "条目名称" in row_text and ("条目编码" in row_text or "一级分类" in row_text):
            header_row = r
            break
    if not header_row:
        header_row = 3  # 兜底

    # 2) 建立列名 -> 列号 映射
    colmap = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v and str(v).strip() in FIELD_MAP:
            colmap[FIELD_MAP[str(v).strip()]] = c

    def col(field):
        return colmap.get(field, FIELD_COL_FALLBACK.get(field, 1))

    items = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, col("name")).value
        if not name:
            continue

        desc = ws.cell(r, col("desc")).value or ""
        short_name, auto_desc = split_title_desc(str(name))
        if not desc:
            desc = auto_desc

        it = {
            "code": ws.cell(r, col("code")).value or "",
            "system": ws.cell(r, col("system")).value or "",
            "cat1": ws.cell(r, col("cat1")).value or "",
            "cat2": ws.cell(r, col("cat2")).value or "",
            "area": ws.cell(r, col("area")).value or "",
            "unit": ws.cell(r, col("unit")).value or "",
            "name": str(name).strip(),
            "short": short_name,
            "desc": str(desc).strip(),
            "method": ws.cell(r, col("method")).value or "",
            "score_lo": ws.cell(r, col("score_lo")).value,
            "score_hi": ws.cell(r, col("score_hi")).value,
            "severity": ws.cell(r, col("severity")).value or "",
        }
        items.append(it)

    cat1_list = sorted({it["cat1"] for it in items if it["cat1"]})
    cat2_list = sorted({it["cat2"] for it in items if it["cat2"]})
    return items, cat1_list, cat2_list


def make_template_bytes(items=None, title=None):
    """生成导入模板（或导出已有标准），返回 xlsx 字节。

    items: 可选，已有条目列表（字段名同 FIELD_MAP 的值，如 idx/code/cat1...）。
           提供则模板带真实内容（可作示例 / 也可作为「转换后」的导出文件）；
           不提供则写入 2 行示例，避免空白表格让人误以为下载失败。
    title: 可选，A1 单元格的提示文字。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "检查标准条目"
    if title:
        ws["A1"] = title
    headers = list(FIELD_MAP.keys())  # 序号, 条目编码, ...
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(2, i, h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DDEBF7")

    if items:
        r = 3
        for it in items[:300]:
            if not isinstance(it, dict):
                continue
            for i, h in enumerate(headers, start=1):
                f = FIELD_MAP[h]          # 列名 -> 字段名
                v = it.get(f)
                # 兼容：有些来源用 'name' 但 descless 时把描述拼在 name 里，这里原样写出
                ws.cell(r, i, v)
            r += 1
    else:
        # 2 行示例，避免空表
        demo = [
            {"idx": 1, "code": "DEMO001", "system": "示例体系", "cat1": "消防", "cat2": "灭火器",
             "area": "罐区", "unit": "", "name": "（示例）灭火器压力不足。压力表指针低于绿区。",
             "desc": "现场抽查灭火器，压力表不在绿区", "method": "目视检查压力表",
             "score_lo": 1, "score_hi": 5, "severity": "一般问题"},
            {"idx": 2, "code": "DEMO002", "system": "示例体系", "cat1": "安全", "cat2": "用电",
             "area": "配电间", "unit": "", "name": "（示例）配电箱未上锁。非专业人员可随意打开。",
             "desc": "配电箱未锁闭", "method": "现场检查",
             "score_lo": 1, "score_hi": 5, "severity": "重要问题"},
        ]
        r = 3
        for it in demo:
            for i, h in enumerate(headers, start=1):
                ws.cell(r, i, it.get(FIELD_MAP[h]))
            r += 1

    # 适当加宽列，便于阅读
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16
    ws.column_dimensions["H"].width = 50   # 条目名称
    ws.column_dimensions["I"].width = 40   # 问题描述

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


if __name__ == "__main__":
    import sys
    from pathlib import Path
    p = sys.argv[1] if len(sys.argv) > 1 else None
    if p:
        its, c1, c2 = parse_xlsx_to_items(p)
        print(f"解析 {Path(p).name}: {len(its)} 条, 一级分类 {len(c1)} 个")
        for it in its[:3]:
            print("  ", it["code"], it["cat1"], "/", it["cat2"], "->", it["short"])
    else:
        b = make_template_bytes()
        print(f"模板字节数: {len(b)}")
